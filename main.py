from fastapi import FastAPI, UploadFile, File, Request
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pathlib import Path
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import pytesseract
import os
import io


from rag import add_document, search, load
from gemini import ask_gemini


app = FastAPI()

load()

app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="templates")

@app.get("/", response_class=HTMLResponse)
def chat_page(request: Request):

    html = templates.get_template("chat.html").render({"request": request})
    return HTMLResponse(content=html, media_type="text/html; charset=utf-8")

@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request):
    return templates.TemplateResponse("admin.html", {"request": request})



def extract_text_from_upload(file: UploadFile, data: bytes) -> tuple[str, str]:
    """
    Returns: (text, error_message). error_message is "" if ok.
    """
    filename = (file.filename or "").lower()

    # 1) TXT
    if filename.endswith(".txt"):
        return data.decode("utf-8", errors="ignore").strip(), ""

    # 2) PDF
    if filename.endswith(".pdf"):
        reader = PdfReader(io.BytesIO(data))
        pages = [(p.extract_text() or "") for p in reader.pages]
        text = "\n".join(pages).strip()
        if not text:
            return "", "PDF has no extractable text (maybe scanned). Use OCR or upload a text-based PDF."
        return text, ""

    # 3) DOCX
    if filename.endswith(".docx"):
        doc = Document(io.BytesIO(data))
        text = "\n".join([p.text for p in doc.paragraphs]).strip()
        if not text:
            return "", "DOCX has no readable text."
        return text, ""

    # 4) Excel
    if filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(data), data_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"--- Sheet: {ws.title} ---")
            for row in ws.iter_rows(values_only=True):
                row_vals = [str(v) for v in row if v is not None and str(v).strip() != ""]
                if row_vals:
                    lines.append(" | ".join(row_vals))
        text = "\n".join(lines).strip()
        if not text:
            return "", "Excel file seems empty."
        return text, ""

    # 5) Images (OCR)
    if filename.endswith((".png", ".jpg", ".jpeg", ".webp")):
        img = Image.open(io.BytesIO(data))
        text = pytesseract.image_to_string(img).strip()
        if not text:
            return "", "No text detected in the image."
        return text, ""
        
    return "", "Unsupported file type. Supported: .txt, .pdf, .docx, .xlsx, .png, .jpg, .jpeg, .webp"

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    data = await file.read()

    # 1) Save original file to uploads/
    safe_name = os.path.basename(file.filename or "uploaded_file")
    file_path = UPLOAD_DIR / safe_name
    with open(file_path, "wb") as f:
        f.write(data)

    text, err = extract_text_from_upload(file, data)
    if err:
        return JSONResponse({"message": err}, status_code=400)

    # IMPORTANT: Large files should be chunked (see section 3)
    add_document(text)

    return {"message": f"Uploaded, saved, and indexed: {safe_name}"}



@app.post("/ask")
async def ask(request: Request):

    data = await request.json()

    question = data["question"]

    try:
        context = search(question)
        if not context:
            return {"answer": "I couldn't find any relevant information in the documents."}
        
        answer = ask_gemini(context, question)

        return {"answer": answer}
    except Exception as e:
        return {"answer": f"An error occurred: {str(e)}"}
    
@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return FileResponse("static/favicon.ico")